from __future__ import annotations
import io
import uuid
from datetime import datetime, date
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.models import Article, ScopedPricingRule, BulkUpload

router = APIRouter(prefix="/bulk", tags=["bulk"])

# ── Template column definitions ──────────────────────────────────────────────
TEMPLATE_HEADERS = [
    "Serviceability Level",         # National | Region | State | Merchant Cohort | Distributor
    "Serviceability Value",         # ALL (National) | region code | state code | cohort/dist ID
    "Merchant Classification Type", # Distributor-Retailer | Distributor-Wholesaler | Anchor Distributor
    "Segment Type",                 # Beverages | FMCG | Staples | etc. (optional)
    "Category L1",                  # optional
    "Category L2",                  # optional
    "Category L3",                  # optional
    "Brand ID",                     # optional
    "Article ID",                   # mandatory SKU/article code
    "Article Group ID",             # optional virtual group
    "MRP",                          # mandatory, > 0
    "Retailer Margin Type",         # Markdown as % of MRP | Re off on MRP (absolute) | Absolute PTR
    "Retailer Margin Value",        # numeric value
    "Distributor Margin Type",      # Markdown as % of PTR | Re off on PTR (absolute) | Absolute PTD
    "Distributor Margin Value",     # numeric value
    "Start Date",                   # DD-MM-YYYY
    "End Date",                     # DD-MM-YYYY
]

SCOPE_LEVEL_MAP = {
    "national":       "NATIONAL",
    "region":         "NATIONAL_DPG",
    "state":          "STATE_DPG",
    "merchant cohort": "CUSTOMER",
    "distributor":    "CUSTOMER",
}

SCOPE_PRIORITY_MAP = {
    "CUSTOMER":     1,
    "STATE_DPG":    2,
    "NATIONAL_DPG": 3,
    "NATIONAL":     4,
}

RM_TYPE_MAP = {
    "markdown as % of mrp":          "PCT_MRP",
    "re off on mrp (absolute)":      "ABS",
    "absolute ptr":                   "ABS_PTR",
    # aliases from the RCPL xlsx
    "cost per each - re off on mrp":  "ABS",
    "cost per case - re off on mrp":  "ABS",
    "absolute value":                 "ABS_PTR",
}

DM_TYPE_MAP = {
    "markdown as % of ptr":          "PCT_PTR",
    "re off on ptr (absolute)":      "ABS",
    "absolute ptd":                   "ABS_PTD",
    # common aliases
    "cost per each - re off on ptr":  "ABS",
    "cost per case - re off on ptr":  "ABS",
    "absolute value":                 "ABS_PTD",
}

SAMPLE_ROW = [
    "National", "ALL", "Distributor-Retailer", "Beverages",
    "", "", "", "",
    "SKU_001", "", 100,
    "Markdown as % of MRP", 7,
    "Markdown as % of PTR", 3,
    "01-06-2026", "31-12-2026",
]


@router.get("/template")
async def download_template():
    wb = Workbook()

    # ── Template sheet ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Pricing Template"

    header_fill = PatternFill("solid", fgColor="1B3FA6")
    optional_fill = PatternFill("solid", fgColor="3B5FD4")
    mandatory_cols = {0, 1, 2, 8, 10, 11, 12, 13, 14}   # 0-indexed

    for i, header in enumerate(TEMPLATE_HEADERS):
        cell = ws.cell(row=1, column=i + 1, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = optional_fill if i not in mandatory_cols else header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # sample row
    for i, val in enumerate(SAMPLE_ROW):
        ws.cell(row=2, column=i + 1, value=val)

    col_widths = [22, 22, 28, 18, 14, 14, 14, 12, 16, 18, 10, 30, 18, 30, 18, 14, 14]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[ws.cell(1, i + 1).column_letter].width = w
    ws.row_dimensions[1].height = 36

    # ── Attribute Definition sheet ──────────────────────────────────────────
    info = wb.create_sheet("Attribute Definition")
    info_headers = ["Column", "Mandatory", "Allowed Values / Notes"]
    attr_rows = [
        ("Serviceability Level",         "Y", "National | Region | State | Merchant Cohort | Distributor"),
        ("Serviceability Value",          "Y", "Use 'ALL' for National. Region code, state code, cohort/distributor ID otherwise."),
        ("Merchant Classification Type",  "Y", "Distributor-Retailer | Distributor-Wholesaler | Anchor Distributor"),
        ("Segment Type",                  "N", "Beverages | FMCG | Staples | etc."),
        ("Category L1",                   "N", "Category level 1"),
        ("Category L2",                   "N", "Category level 2"),
        ("Category L3",                   "N", "Category level 3"),
        ("Brand ID",                      "N", "Brand identifier"),
        ("Article ID",                    "Y", "Internal SKU/article code from MDM"),
        ("Article Group ID",              "N", "Virtual article group ID (from Virtual Grouping file)"),
        ("MRP",                           "Y", "Maximum Retail Price, numeric > 0"),
        ("Retailer Margin Type",          "Y", "Markdown as % of MRP | Re off on MRP (absolute) | Absolute PTR"),
        ("Retailer Margin Value",         "Y", "Numeric. % value for Markdown, absolute amount for others."),
        ("Distributor Margin Type",       "Y", "Markdown as % of PTR | Re off on PTR (absolute) | Absolute PTD"),
        ("Distributor Margin Value",      "Y", "Numeric. % value for Markdown, absolute amount for others."),
        ("Start Date",                    "Y", "DD-MM-YYYY format"),
        ("End Date",                      "Y", "DD-MM-YYYY format"),
    ]
    for i, h in enumerate(info_headers):
        c = info.cell(row=1, column=i + 1, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
    for r, row in enumerate(attr_rows, 2):
        for c, val in enumerate(row, 1):
            info.cell(row=r, column=c, value=val)
    info.column_dimensions["A"].width = 32
    info.column_dimensions["B"].width = 12
    info.column_dimensions["C"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rcpl_retailer_pricing_template.xlsx"},
    )


def _parse_date(val) -> date | None:
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_row(row_values: list, row_num: int) -> tuple[dict | None, str | None]:
    def cell(idx):
        if idx < len(row_values) and row_values[idx] is not None:
            return str(row_values[idx]).strip()
        return ""

    def num(idx):
        v = row_values[idx] if idx < len(row_values) else None
        if v is None or str(v).strip() == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    scope_level_raw  = cell(0)
    scope_value      = cell(1) or "ALL"
    merch_type       = cell(2)
    segment_type     = cell(3)
    cat_l1           = cell(4)
    cat_l2           = cell(5)
    cat_l3           = cell(6)
    brand_id         = cell(7)
    article_id       = cell(8)
    article_group    = cell(9)
    mrp_val          = num(10)
    rm_type_raw      = cell(11)
    rm_value         = num(12)
    dm_type_raw      = cell(13)
    dm_value         = num(14)
    start_date       = cell(15)
    end_date         = cell(16)

    if not scope_level_raw:
        return None, "Serviceability Level is empty"
    if not article_id and not article_group:
        return None, "Article ID and Article Group ID are both empty — at least one is required"
    if mrp_val is None or mrp_val <= 0:
        return None, f"Invalid or missing MRP: {row_values[10] if len(row_values) > 10 else 'N/A'}"
    if not rm_type_raw:
        return None, "Retailer Margin Type is empty"
    if rm_value is None:
        return None, "Retailer Margin Value is empty or invalid"
    if not dm_type_raw:
        return None, "Distributor Margin Type is empty"
    if dm_value is None:
        return None, "Distributor Margin Value is empty or invalid"

    scope_level = SCOPE_LEVEL_MAP.get(scope_level_raw.lower())
    if not scope_level:
        return None, f"Unknown Serviceability Level '{scope_level_raw}'. Use: National, Region, State, Merchant Cohort, Distributor"

    rm_type = RM_TYPE_MAP.get(rm_type_raw.lower())
    if not rm_type:
        return None, (
            f"Unknown Retailer Margin Type '{rm_type_raw}'. "
            "Use: Markdown as % of MRP | Re off on MRP (absolute) | Absolute PTR"
        )

    dm_type = DM_TYPE_MAP.get(dm_type_raw.lower())
    if not dm_type:
        return None, (
            f"Unknown Distributor Margin Type '{dm_type_raw}'. "
            "Use: Markdown as % of PTR | Re off on PTR (absolute) | Absolute PTD"
        )

    return {
        "scope_level":    scope_level,
        "scope_value":    scope_value,
        "customer_group": merch_type or "All",
        "segment_type":   segment_type,
        "cat_l1": cat_l1, "cat_l2": cat_l2, "cat_l3": cat_l3,
        "brand_id":       brand_id,
        "article_id":     article_id,
        "article_group":  article_group,
        "mrp":            mrp_val,
        "rm_type":        rm_type,
        "rm_value":       rm_value,
        "dm_type":        dm_type,
        "dm_value":       dm_value,
        "start_date":     start_date,
        "end_date":       end_date,
    }, None


def _build_scoped_rule_fields(data: dict) -> dict:
    """Convert parsed row into ScopedPricingRule field values."""
    mrp = data["mrp"]

    # ── Retailer margin (MRP → PTR) ──
    rm1 = 0.0
    absolute_ptr = None
    if data["rm_type"] == "PCT_MRP":
        rm1 = round(mrp * data["rm_value"] / 100, 2)
    elif data["rm_type"] == "ABS":
        rm1 = data["rm_value"]
    elif data["rm_type"] == "ABS_PTR":
        absolute_ptr = data["rm_value"]

    # ── Distributor margin (PTR → PTD) ──
    # PTR is estimated here for percentage-based DM calculation
    ptr = absolute_ptr if absolute_ptr is not None else (mrp - rm1)
    dm1 = 0.0
    absolute_ptd = None
    if data["dm_type"] == "PCT_PTR":
        dm1 = round(ptr * data["dm_value"] / 100, 2)
    elif data["dm_type"] == "ABS":
        dm1 = data["dm_value"]
    elif data["dm_type"] == "ABS_PTD":
        absolute_ptd = data["dm_value"]

    return {
        "scope_level":    data["scope_level"],
        "scope_value":    data["scope_value"],
        "customer_group": data["customer_group"],
        "mrp":            mrp,
        "priority":       SCOPE_PRIORITY_MAP.get(data["scope_level"], 4),
        "rm1": rm1, "rm2": 0,
        "absolute_ptr":   absolute_ptr,
        "dm1": dm1, "dm2": 0,
        "absolute_ptd":   absolute_ptd,
        "valid_from": _parse_date(data["start_date"]),
        "valid_to":   _parse_date(data["end_date"]),
    }


@router.post("/upload")
async def upload_pricing(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx or .xls files are accepted")

    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Could not parse Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows = [r for r in rows if any(c is not None for c in r)]

    upload = BulkUpload(filename=file.filename, status="PROCESSING", total_rows=len(rows))
    db.add(upload)
    await db.flush()

    if not rows:
        upload.status = "FAILED"
        upload.failed_count = 0
        upload.error_details = {"errors": [{"row": 0, "error": "File has no data rows"}]}
        upload.completed_at = datetime.utcnow()
        await db.commit()
        return _upload_out(upload)

    errors = []
    success = 0

    for i, row in enumerate(rows):
        row_num = i + 2
        data, err = _parse_row(list(row), row_num)
        if err:
            errors.append({"row": row_num, "article_id": str(row[8] or ""), "error": err})
            continue

        # Look up article by Article ID (SKU) or skip if only group ID given
        article = None
        if data["article_id"]:
            result = await db.execute(select(Article).where(Article.sku == data["article_id"]))
            article = result.scalar_one_or_none()
            if not article:
                errors.append({"row": row_num, "article_id": data["article_id"],
                               "error": "Article ID not found — create the article first"})
                continue

        if not article:
            errors.append({"row": row_num, "article_id": data.get("article_group", ""),
                           "error": "Article Group upload not yet supported — provide Article ID"})
            continue

        fields = _build_scoped_rule_fields(data)
        rule = ScopedPricingRule(article_id=article.id, **fields)
        db.add(rule)
        success += 1

    upload.success_count = success
    upload.failed_count = len(errors)
    upload.error_details = {"errors": errors} if errors else None
    upload.completed_at = datetime.utcnow()
    upload.status = "SUCCESS" if not errors else ("PARTIAL" if success > 0 else "FAILED")

    await db.commit()
    return _upload_out(upload)


@router.get("/uploads")
async def list_uploads(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(BulkUpload).order_by(BulkUpload.created_at.desc()).limit(50))
    return [_upload_out(u) for u in result.scalars().all()]


@router.get("/uploads/{upload_id}")
async def get_upload(upload_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(BulkUpload).where(BulkUpload.id == upload_id))
    upload = result.scalar_one_or_none()
    if not upload:
        raise HTTPException(404, "Upload not found")
    return _upload_out(upload)


def _upload_out(u: BulkUpload) -> dict:
    return {
        "id": str(u.id),
        "filename": u.filename,
        "status": u.status,
        "total_rows": u.total_rows,
        "success_count": u.success_count,
        "failed_count": u.failed_count,
        "error_details": u.error_details,
        "created_by": u.created_by or "System User",
        "created_at": u.created_at.isoformat() if u.created_at else None,
        "completed_at": u.completed_at.isoformat() if u.completed_at else None,
    }
